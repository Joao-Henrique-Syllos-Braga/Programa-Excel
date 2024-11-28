from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

inf = False

def create_excel(data, fileName, inf):
    # Creating a new workbook
    wb = Workbook()

    # Accessing the active sheet
    sheet = wb.active
    sheet.title = "Listão"

    # Defining the background color and font color styles
    fill = PatternFill(start_color="1f1f1f", end_color="1f1f1f", fill_type="solid")
    font = Font(color="ffffff")
    alignment = Alignment(horizontal="center", vertical="center")

    # Adding the first row with styles
    first_row = data[0]  # First row of data
    
    # Apply styles to each cell in the first row
    for col_num, value in enumerate(first_row, 1):  # Start counting from 1 for Excel columns
        cell = sheet.cell(row=1, column=col_num, value=value)
        cell.alignment = alignment  # Apply the alignment
        cell.fill = fill            # Apply the background color
        cell.font = font            # Apply the font color

    # Add the remaining rows (data starting from the second row)
    for row in data[1:]:
        # Ensure that each value in the row is a single string or number (not a list)
        processed_row = []
        for value in row:
            if isinstance(value, list):  # If it's a list, join the characters into a single string
                processed_row.append("".join(value))  # Join the characters into a single string
            else:
                processed_row.append(value)  # Otherwise, leave the value as it is
        sheet.append(processed_row)

    # Define column width (e.g., columns for "Descrição" and "Fornecedor")
    sheet.column_dimensions["D"].width = 20  # Example column, you can customize this
    sheet.column_dimensions["E"].width = 20  # Example column, you can customize this
    sheet.column_dimensions["K"].width = 20  # Example column, you can customize this
    sheet.column_dimensions["L"].width = 20  # Example column, you can customize this
    sheet.column_dimensions["H"].width = 10  # Example column, you can customize this
    sheet.column_dimensions["C"].width = 10  # Example column, you can customize this
    sheet.column_dimensions["F"].width = 10  # Example column, you can customize this

    # Saving the file
    if inf:
        wb.save(fileName)
